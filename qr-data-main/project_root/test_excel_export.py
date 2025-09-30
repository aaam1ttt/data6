#!/usr/bin/env python3
"""
Test Excel export with auto-adjusted column widths
"""

import sys
import os
import io
sys.path.append('.')

from app import create_app
from app.routes.scan import _auto_adjust_column_widths
from app.core.forms_parser import (
    torg12_make_string, env_make_string, exploitation_make_string,
    transport_make_string, custom_make_string, TORG12_FIELDS,
    torg12_parse_string, env_parse_string, exploitation_parse_string,
    transport_parse_string, custom_parse_string
)
import openpyxl

def test_excel_export():
    """Test Excel export with auto-adjusted column widths"""
    
    print("Testing Excel export with auto-adjusted column widths...")
    
    # Test 1: TORG-12
    print("\n1. Testing TORG-12 export...")
    values = {
        "01": "ООО 'Тест Компания' с очень длинным названием организации",
        "02": "12345678",
        "03": "г. Москва, ул. Тестовая, д. 123, корп. 4, офис 567",
        "04": "+7 (495) 123-45-67"
    }
    text = torg12_make_string(values)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ТОРГ-12"
    ws.append(["Код", "Наименование", "Значение"])
    parsed = torg12_parse_string(text)
    for code, label in TORG12_FIELDS[:10]:  # First 10 rows for testing
        ws.append([code, label, parsed.get(code, "")])
    
    _auto_adjust_column_widths(ws)
    
    # Check column widths
    for col in ['A', 'B', 'C']:
        width = ws.column_dimensions[col].width
        print(f"  Column {col} width: {width}")
        assert width > 0, f"Column {col} width should be > 0"
    
    # Column B and C should be wider due to longer content
    assert ws.column_dimensions['B'].width > 10, "Column B should be wider"
    assert ws.column_dimensions['C'].width > 10, "Column C should be wider"
    
    print("  [OK] TORG-12 column widths adjusted correctly")
    
    # Test 2: Message
    print("\n2. Testing Message export...")
    pairs = [
        ("Параметр с длинным названием", "Значение"),
        ("Параметр", "Очень длинное значение для проверки автоматической настройки ширины")
    ]
    text = env_make_string(pairs)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сообщение"
    ws.append(["Параметр", "Значение"])
    for p, v in env_parse_string(text):
        ws.append([p, v])
    
    _auto_adjust_column_widths(ws)
    
    for col in ['A', 'B']:
        width = ws.column_dimensions[col].width
        print(f"  Column {col} width: {width}")
        assert width > 0, f"Column {col} width should be > 0"
    
    print("  [OK] Message column widths adjusted correctly")
    
    # Test 3: Exploitation
    print("\n3. Testing Exploitation export...")
    rows = [
        ("Обозначение 1", "Входимость 1", "Носитель длинный", "SN123456789", "UUID-12345-67890-ABCDEF"),
        ("Обозн. 2", "Вход. 2", "Нос.", "SN987", "UUID-98765")
    ]
    text = exploitation_make_string(rows)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Эксплуатация"
    ws.append(["Обозначение СЧ","Входимость СЧ","Носитель маркировки","Серийный номер","Уникальный идентификатор"])
    for row in exploitation_parse_string(text):
        ws.append(row)
    
    _auto_adjust_column_widths(ws)
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        width = ws.column_dimensions[col].width
        print(f"  Column {col} width: {width}")
        assert width > 0, f"Column {col} width should be > 0"
    
    print("  [OK] Exploitation column widths adjusted correctly")
    
    # Test 4: Transport
    print("\n4. Testing Transport export...")
    rows = [
        ("Знак 1", "Значение с очень длинным текстом", "Вид данных", "12345"),
        ("Знак 2", "Знач.", "Вид", "67890")
    ]
    text = transport_make_string(rows)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Транспорт"
    ws.append(["Номер знака","Значение","Вид данных","Цифровое значение"])
    for row in transport_parse_string(text):
        ws.append(row)
    
    _auto_adjust_column_widths(ws)
    
    for col in ['A', 'B', 'C', 'D']:
        width = ws.column_dimensions[col].width
        print(f"  Column {col} width: {width}")
        assert width > 0, f"Column {col} width should be > 0"
    
    print("  [OK] Transport column widths adjusted correctly")
    
    # Test 5: Custom table
    print("\n5. Testing Custom table export...")
    rows = [
        ["Короткий", "Средний текст", "Очень длинный текст для проверки автоматической настройки"],
        ["A", "B", "C"],
        ["123", "456789", "0"]
    ]
    text = custom_make_string(rows)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Таблица"
    for row in custom_parse_string(text):
        ws.append(row)
    
    _auto_adjust_column_widths(ws)
    
    for col in ['A', 'B', 'C']:
        width = ws.column_dimensions[col].width
        print(f"  Column {col} width: {width}")
        assert width > 0, f"Column {col} width should be > 0"
    
    # Column C should be widest
    assert ws.column_dimensions['C'].width > ws.column_dimensions['A'].width, \
        "Column C should be wider than A"
    
    print("  [OK] Custom table column widths adjusted correctly")
    
    # Test 6: Save to file and verify
    print("\n6. Testing file save...")
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # Reload and verify
    wb_reloaded = openpyxl.load_workbook(bio)
    ws_reloaded = wb_reloaded.active
    
    for col in ['A', 'B', 'C']:
        width = ws_reloaded.column_dimensions[col].width
        assert width > 0, f"Reloaded column {col} width should be > 0"
    
    print("  [OK] Excel file saved and loaded successfully with adjusted widths")
    
    print("\n[SUCCESS] All Excel export tests PASSED!")
    return True

if __name__ == "__main__":
    try:
        test_excel_export()
    except Exception as e:
        print(f"\n[FAILED] Test failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
